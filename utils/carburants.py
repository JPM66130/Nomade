import json
import io
import urllib.request
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

SOURCE_URL = "https://donnees.roulez-eco.fr/opendata/instantane"
EU_SOURCE_URL = "https://energy.ec.europa.eu/document/download/264c2d0f-f161-4ea3-a777-78faae59bea0_en?filename=Weekly%20Oil%20Bulletin%20Weekly%20prices%20with%20Taxes%20-%202024-02-19.xlsx"
CACHE_PATH = Path(__file__).resolve().parent.parent / "fuel_prices_cache.json"
CACHE_TTL = timedelta(hours=24)
EU_CACHE_PATH = Path(__file__).resolve().parent.parent / "fuel_prices_eu_cache.json"
EU_CACHE_TTL = timedelta(days=7)


def _read_prices(xml_bytes):
    root = ElementTree.fromstring(xml_bytes)
    totals = {}
    counts = {}
    for price in root.findall(".//prix"):
        name = (price.attrib.get("nom") or "").strip().lower()
        value = price.attrib.get("valeur")
        if not name or not value:
            continue
        try:
            amount = float(value.replace(",", "."))
        except ValueError:
            continue
        totals[name] = totals.get(name, 0.0) + amount
        counts[name] = counts.get(name, 0) + 1
    return {name: round(totals[name] / counts[name], 3) for name in totals}


def _load_cache():
    if not CACHE_PATH.is_file():
        return None
    try:
        data = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        updated_at = datetime.fromisoformat(data["updated_at"])
        if datetime.now(timezone.utc) - updated_at < CACHE_TTL:
            return data
    except (KeyError, ValueError, OSError, json.JSONDecodeError):
        return None
    return None


def get_fuel_prices():
    cached = _load_cache()
    if cached:
        return cached
    try:
        request = urllib.request.Request(SOURCE_URL, headers={"User-Agent": "Itineraire-C25/1.0"})
        with urllib.request.urlopen(request, timeout=20) as response:
            archive = response.read()
        with zipfile.ZipFile(io.BytesIO(archive)) as archive_file:
            xml_name = next(name for name in archive_file.namelist() if name.lower().endswith(".xml"))
            prices = _read_prices(archive_file.read(xml_name))
        data = {
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "source": SOURCE_URL,
            "prices_eur_l": prices,
            "stale": False,
        }
        temporary = CACHE_PATH.with_suffix(".tmp")
        temporary.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        temporary.replace(CACHE_PATH)
        return data
    except Exception:
        return _stale_cache()


def get_eu_fuel_prices():
    cached = _load_generic_cache(EU_CACHE_PATH, EU_CACHE_TTL)
    if cached:
        return cached
    try:
        request = urllib.request.Request(EU_SOURCE_URL, headers={"User-Agent": "Itineraire-C25/1.0"})
        with urllib.request.urlopen(request, timeout=20) as response:
            workbook = load_workbook(io.BytesIO(response.read()), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header_index = 0
        country_index = 0
        diesel_index = 2
        prices = {}
        for row in rows[header_index + 1:]:
            if row[country_index] and isinstance(row[diesel_index], (int, float)):
                prices[str(row[country_index]).strip().lower()] = round(float(row[diesel_index]) / 1000, 3)
        return _write_cache(EU_CACHE_PATH, EU_SOURCE_URL, prices, EU_CACHE_TTL)
    except Exception:
        return _stale_cache(EU_CACHE_PATH, EU_SOURCE_URL)


def _load_generic_cache(path, ttl):
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        updated_at = datetime.fromisoformat(data["updated_at"])
        if datetime.now(timezone.utc) - updated_at < ttl:
            return data
    except (KeyError, ValueError, OSError, json.JSONDecodeError):
        return None
    return None


def _write_cache(path, source, prices, ttl):
    data = {"updated_at": datetime.now(timezone.utc).isoformat(), "source": source, "prices_eur_l": prices, "stale": False}
    temporary = path.with_suffix(".tmp")
    temporary.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    temporary.replace(path)
    return data


def _stale_cache(path=CACHE_PATH, source=SOURCE_URL):
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        data["stale"] = True
        return data
    except (OSError, json.JSONDecodeError):
        return {"updated_at": None, "source": source, "prices_eur_l": {}, "stale": True}
